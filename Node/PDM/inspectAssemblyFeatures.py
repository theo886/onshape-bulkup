#!/usr/bin/env python3
"""
inspectAssemblyFeatures.py — Detect assembly-level geometry features in SolidWorks assemblies
and export affected assemblies as .SLDPRT (exterior faces) for clean Onshape import.

Onshape drops assembly-level geometry (cuts, extrudes, fillets, etc.) on import — only mates
and component instances come through. This script finds assemblies with those features and
exports them as flat part files that preserve the correct visual geometry.

Usage:
    python inspectAssemblyFeatures.py -i Upload/Onshape_Upload_List.xlsx -o output/assembly_inspect
    python inspectAssemblyFeatures.py -i Upload/Onshape_Upload_List.xlsx --dry-run -v

Requires: SolidWorks installed, pywin32, openpyxl
"""

import argparse
import json
import logging
import os
import sys
import time
from datetime import datetime

import openpyxl

# ---------------------------------------------------------------------------
# Feature type allowlist — anything NOT here is flagged as geometry-modifying
# ---------------------------------------------------------------------------
ALLOWED_FEATURE_TYPES = frozenset({
    # Mates
    "MateGroup", "MateCoincident4", "MateParallel", "MatePerpendicular",
    "MateTangent", "MateConcentric4", "MateDistanceDim", "MateAngleDim",
    "MateLock", "MateWidth", "MateGear", "MateRackPinion", "MateScrew",
    "MateCamFollower", "MateLinearCoupler", "MateSlot", "MateHinge",
    "MateUniversalJoint", "MatePathMate", "MateSymmetric",
    "MateProfileCenter", "MateMagneticMate", "MateLimit", "SmartMate",
    "MatePlaneInPlace",
    # Reference geometry
    "RefPlane", "RefAxis", "RefPoint", "CoordSys", "Reference",
    "OriginProfileFeature", "BOM_CSYS",
    # Components
    "ComponentPattern", "ComponentPattern2", "AsmPattern",
    "SmartComponents", "CompFeatureData",
    # Structure / folders
    "Folder", "FtrFolder", "SubFolder", "SubAtomFolder",
    "CommentsFolder", "FlattenFolder",
    # Annotations
    "Comment", "Attribute", "Sensor", "SensorFolder", "DOFManager",
    # Display
    "Exploded", "ExplodedStep", "ExplodeLineSketch",
    "SelectionSet", "SelectionSet2",
    # In-context editing / motion
    "AsmEdit", "DriveComponent",
    # Layout
    "LayoutSketch", "AsmLayout",
    # Origin / system
    "HistoryFolder", "DetailCabinet", "ProfileFeature", "LiveSection",
})

# SolidWorks constants
SW_DOC_ASSEMBLY = 2
SW_OPEN_DOC_SILENT = 1
SW_SAVE_AS_COPY = 1
SW_SAVE_AS_SILENT = 2
SW_SAVE_ASSEMBLY_AS_PART_OPTIONS_PREF = 65  # swUserPreferenceIntegerValue_e


# ---------------------------------------------------------------------------
# SolidWorks COM helpers
# ---------------------------------------------------------------------------

def connect_solidworks():
    """Connect to a running SolidWorks instance, or launch one."""
    import win32com.client

    try:
        sw_app = win32com.client.GetActiveObject("SldWorks.Application")
        logging.info("Connected to running SolidWorks instance")
    except Exception:
        sw_app = win32com.client.Dispatch("SldWorks.Application")
        sw_app.Visible = False
        logging.info("Launched new SolidWorks instance (hidden)")
    return sw_app


def open_assembly(sw_app, file_path):
    """Open an assembly in SolidWorks (silent mode). Returns (doc, errors, warnings).

    Under pywin32 dynamic dispatch, byref out-params are unreliable when passed as
    VARIANT objects. Passing plain integers causes pywin32 to return all outputs
    (including byref params) as a tuple instead.
    """
    result = sw_app.OpenDoc6(
        file_path,
        SW_DOC_ASSEMBLY,
        SW_OPEN_DOC_SILENT,
        "",
        0,  # errors (byref out-param — returned in tuple)
        0,  # warnings (byref out-param — returned in tuple)
    )
    # pywin32 dynamic dispatch returns (doc, errors, warnings) tuple
    if isinstance(result, tuple):
        return result[0], result[1], result[2]
    # Fallback: result is just the doc (some pywin32 versions)
    return result, 0, 0


def close_document(sw_app, doc):
    """Close a document safely."""
    if doc is not None:
        try:
            title = doc.GetTitle()
            sw_app.CloseDoc(title)
        except Exception as e:
            logging.warning("Failed to close document: %s", e)


# ---------------------------------------------------------------------------
# Feature inspection
# ---------------------------------------------------------------------------

def inspect_features(doc, verbose=False):
    """
    Walk the feature tree and return info about non-allowlisted, unsuppressed features.

    Returns dict with:
        flagged_features: list of (name, type_name) tuples
        all_types_encountered: set of all feature type names seen
        total_features: int count of all features walked
    """
    flagged = []
    all_types = set()
    total = 0

    # Get active configuration name for IsSuppressed2 check
    active_config = doc.GetActiveConfiguration()
    config_name = active_config.Name if active_config else ""

    feature = doc.FirstFeature()
    while feature is not None:
        type_name = feature.GetTypeName2()
        all_types.add(type_name)
        total += 1

        # Skip suppressed features — only inspect active (unsuppressed) ones.
        # IsSuppressed2(nNumConfigs, configNameArray) returns an array of booleans,
        # one per config. We check just the active configuration.
        suppressed = False
        try:
            is_suppressed = feature.IsSuppressed2(1, [config_name])
            if is_suppressed is not None:
                if isinstance(is_suppressed, (list, tuple)):
                    suppressed = bool(is_suppressed[0])
                else:
                    suppressed = bool(is_suppressed)
        except Exception:
            # If IsSuppressed2 fails, fall back to the simpler property
            try:
                suppressed = bool(feature.IsSuppressed)
            except Exception:
                suppressed = False

        if not suppressed and type_name not in ALLOWED_FEATURE_TYPES:
            flagged.append((feature.Name, type_name))
            if verbose:
                logging.info("  FLAGGED: %s (%s)", feature.Name, type_name)
        elif verbose:
            status = "suppressed" if suppressed else "ok"
            logging.debug("  %s: %s (%s) [%s]", status, feature.Name, type_name, status)

        feature = feature.GetNextFeature()

    return {
        "flagged_features": flagged,
        "all_types_encountered": all_types,
        "total_features": total,
    }


# ---------------------------------------------------------------------------
# Export assembly as part (exterior faces)
# ---------------------------------------------------------------------------

def export_as_part(sw_app, doc, output_path):
    """
    Save the assembly as a .SLDPRT with exterior faces only.
    Returns (success, error_message).
    """
    # Save original preference
    original_pref = sw_app.GetUserPreferenceIntegerValue(
        SW_SAVE_ASSEMBLY_AS_PART_OPTIONS_PREF
    )

    try:
        # Set to exterior faces mode (0 = ExteriorFaces)
        sw_app.SetUserPreferenceIntegerValue(
            SW_SAVE_ASSEMBLY_AS_PART_OPTIONS_PREF, 0
        )

        # SaveAs3 flags: 1=Copy, 2=Silent → 3
        # Pass plain integers for byref out-params (pywin32 dynamic dispatch pattern)
        result = doc.Extension.SaveAs3(
            output_path,
            0,  # version (0 = current)
            SW_SAVE_AS_COPY | SW_SAVE_AS_SILENT,
            None,  # export data
            None,  # advanced save options
            0,     # errors (byref out-param)
            0,     # warnings (byref out-param)
        )

        # pywin32 may return (success, errors, warnings) tuple or just success bool
        if isinstance(result, tuple):
            success, errors, warnings = result[0], result[1], result[2]
        else:
            success, errors, warnings = result, 0, 0

        if success:
            return True, None
        else:
            return False, f"SaveAs3 failed: errors={errors}, warnings={warnings}"

    except Exception as e:
        return False, str(e)
    finally:
        # Restore original preference
        sw_app.SetUserPreferenceIntegerValue(
            SW_SAVE_ASSEMBLY_AS_PART_OPTIONS_PREF, original_pref
        )


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

def read_excel(input_path):
    """
    Read the migration Excel and filter to uploadLevel >= 2 (assemblies).
    Returns list of dicts with filePath, partNumber, documentName, row index.
    """
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    # Find column indices
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[h.strip()] = i

    # Required columns
    file_path_col = col_map.get("filePath")
    upload_level_col = col_map.get("uploadLevel")

    if file_path_col is None:
        raise ValueError("Excel missing required column: filePath")
    if upload_level_col is None:
        raise ValueError("Excel missing required column: uploadLevel")

    # Optional columns
    part_number_col = col_map.get("property:Part Number", col_map.get("Part Number"))
    doc_name_col = col_map.get("document:name")

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cells = list(row)

        # Get upload level
        level_val = cells[upload_level_col].value
        if level_val is None:
            continue
        try:
            level = int(level_val)
        except (ValueError, TypeError):
            continue
        if level < 2:
            continue

        file_path = cells[file_path_col].value
        if not file_path:
            continue

        # Only process .SLDASM files
        if not str(file_path).upper().endswith(".SLDASM"):
            continue

        part_number = None
        if part_number_col is not None and part_number_col < len(cells):
            part_number = cells[part_number_col].value

        doc_name = None
        if doc_name_col is not None and doc_name_col < len(cells):
            doc_name = cells[doc_name_col].value

        # Derive part number from filename if not in Excel
        if not part_number:
            part_number = os.path.splitext(os.path.basename(file_path))[0]

        if not doc_name:
            doc_name = part_number

        rows.append({
            "filePath": os.path.normpath(str(file_path)),
            "partNumber": str(part_number),
            "documentName": str(doc_name),
            "rowIndex": row_idx,
        })

    wb.close()
    return rows


def write_results_excel(results, output_path):
    """Write inspection results to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assembly Feature Inspection"

    columns = [
        "filePath", "partNumber", "documentName", "hasFeatures",
        "featureCount", "featureList", "featureTypes",
        "exportedSldprtPath", "status", "error",
    ]
    ws.append(columns)

    for r in results:
        ws.append([
            r.get("filePath", ""),
            r.get("partNumber", ""),
            r.get("documentName", ""),
            r.get("hasFeatures", ""),
            r.get("featureCount", 0),
            r.get("featureList", ""),
            r.get("featureTypes", ""),
            r.get("exportedSldprtPath", ""),
            r.get("status", ""),
            r.get("error", ""),
        ])

    # Auto-size columns (rough estimate)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

    wb.save(output_path)
    logging.info("Results Excel saved: %s", output_path)


# ---------------------------------------------------------------------------
# JSON sidecar (crash-safe resume)
# ---------------------------------------------------------------------------

def load_progress(sidecar_path):
    """Load progress sidecar. Returns dict keyed by normalized filePath."""
    if os.path.exists(sidecar_path):
        with open(sidecar_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_progress(sidecar_path, progress):
    """Save progress sidecar atomically."""
    tmp_path = sidecar_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(progress, f, indent=2, default=str)
    os.replace(tmp_path, sidecar_path)


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def resolve_output_path(output_dir, part_number, existing_paths):
    """
    Generate a unique output .SLDPRT path, appending _2, _3, etc. for duplicates.
    """
    base_name = f"{part_number}.SLDPRT"
    out_path = os.path.join(output_dir, base_name)
    counter = 2
    while out_path in existing_paths:
        base_name = f"{part_number}_{counter}.SLDPRT"
        out_path = os.path.join(output_dir, base_name)
        counter += 1
    existing_paths.add(out_path)
    return out_path


def process_assemblies(args):
    """Main processing loop."""
    # Read Excel
    logging.info("Reading Excel: %s", args.input)
    rows = read_excel(args.input)
    logging.info("Found %d assemblies (uploadLevel >= 2, .SLDASM)", len(rows))

    if not rows:
        logging.warning("No assemblies found to inspect")
        return

    # Apply --start-index and --count
    start = args.start_index
    if start >= len(rows):
        logging.warning("--start-index %d exceeds row count %d", start, len(rows))
        return
    rows = rows[start:]
    if args.count > 0:
        rows = rows[:args.count]
    logging.info("Processing %d assemblies (start=%d, count=%s)",
                 len(rows), start, args.count if args.count > 0 else "all")

    # Output directory
    os.makedirs(args.output, exist_ok=True)

    # Sidecar
    sidecar_path = args.sidecar or os.path.join(args.output, "inspect_progress.json")
    progress = load_progress(sidecar_path)

    # Track all feature types encountered (for verbose auditing)
    global_all_types = set()

    # Counters
    processed = 0
    succeeded = 0
    has_features_count = 0
    exported = 0
    skipped = 0
    failed = 0

    # Existing output paths (for duplicate detection) — pre-seed from sidecar
    existing_paths = set()
    for entry in progress.values():
        if isinstance(entry, dict) and entry.get("exportedSldprtPath"):
            existing_paths.add(entry["exportedSldprtPath"])

    results = []

    # Connect to SolidWorks (skip in dry-run)
    sw_app = None
    if not args.dry_run:
        try:
            sw_app = connect_solidworks()
        except Exception as e:
            logging.error("Cannot connect to SolidWorks: %s", e)
            logging.error("Make sure SolidWorks is installed and running on Windows")
            sys.exit(1)

    try:
        for idx, row in enumerate(rows):
            file_path = row["filePath"]
            part_number = row["partNumber"]
            doc_name = row["documentName"]

            logging.info("[%d/%d] %s", idx + 1, len(rows), part_number)

            # Check sidecar for resume
            if file_path in progress and not args.force:
                prev = progress[file_path]
                logging.info("  SKIP (already processed): status=%s", prev.get("status"))
                skipped += 1
                results.append(prev)
                continue

            result = {
                "filePath": file_path,
                "partNumber": part_number,
                "documentName": doc_name,
                "hasFeatures": "N",
                "featureCount": 0,
                "featureList": "",
                "featureTypes": "",
                "exportedSldprtPath": "",
                "status": "skipped",
                "error": "",
            }

            # Check file exists
            if not os.path.isfile(file_path):
                result["status"] = "error"
                result["error"] = "File not found"
                logging.error("  File not found: %s", file_path)
                failed += 1
                results.append(result)
                progress[file_path] = result
                save_progress(sidecar_path, progress)
                processed += 1
                continue

            if args.dry_run:
                result["status"] = "dry-run"
                logging.info("  DRY RUN: would inspect %s", file_path)
                results.append(result)
                processed += 1
                continue

            # Open assembly
            doc = None
            try:
                doc, open_errors, open_warnings = open_assembly(sw_app, file_path)
                if doc is None:
                    result["status"] = "error"
                    result["error"] = f"OpenDoc6 failed: errors={open_errors}, warnings={open_warnings}"
                    logging.error("  Failed to open: errors=%d, warnings=%d",
                                  open_errors, open_warnings)
                    failed += 1
                    results.append(result)
                    progress[file_path] = result
                    save_progress(sidecar_path, progress)
                    processed += 1
                    continue

                # Inspect features
                inspection = inspect_features(doc, verbose=args.verbose)
                global_all_types.update(inspection["all_types_encountered"])

                flagged = inspection["flagged_features"]
                if flagged:
                    result["hasFeatures"] = "Y"
                    result["featureCount"] = len(flagged)
                    result["featureList"] = ", ".join(name for name, _ in flagged)
                    result["featureTypes"] = ", ".join(sorted(set(t for _, t in flagged)))
                    has_features_count += 1
                    logging.info("  FOUND %d geometry features: %s",
                                 len(flagged), result["featureTypes"])

                    # Export as part
                    out_path = resolve_output_path(args.output, part_number, existing_paths)
                    success, err_msg = export_as_part(sw_app, doc, out_path)
                    if success:
                        result["exportedSldprtPath"] = out_path
                        result["status"] = "success"
                        exported += 1
                        logging.info("  Exported: %s", out_path)
                    else:
                        result["status"] = "error"
                        result["error"] = f"Export failed: {err_msg}"
                        logging.error("  Export failed: %s", err_msg)
                        failed += 1
                else:
                    result["status"] = "success"
                    logging.info("  Clean (no geometry features)")

                # Only count as succeeded if status is still "success"
                if result["status"] == "success":
                    succeeded += 1

            except Exception as e:
                result["status"] = "error"
                result["error"] = str(e)
                logging.error("  Exception: %s", e)
                failed += 1
            finally:
                close_document(sw_app, doc)
                # Small delay between operations
                time.sleep(0.5)

            results.append(result)
            progress[file_path] = result
            save_progress(sidecar_path, progress)
            processed += 1

    except KeyboardInterrupt:
        logging.warning("\nInterrupted by user — saving progress")
    finally:
        # Save final progress
        save_progress(sidecar_path, progress)

    # Write results Excel
    results_excel_path = os.path.join(args.output, "assembly_feature_inspection.xlsx")
    # Include any previously-processed results from sidecar (filter out __meta__)
    all_results = [v for k, v in progress.items() if k != "__meta__" and isinstance(v, dict)]
    write_results_excel(all_results, results_excel_path)

    # Summary
    logging.info("")
    logging.info("=" * 60)
    logging.info("ASSEMBLY FEATURE INSPECTION COMPLETE")
    logging.info("=" * 60)
    logging.info("Processed:       %d", processed)
    logging.info("Succeeded:       %d", succeeded)
    logging.info("Skipped (resume):%d", skipped)
    logging.info("Failed:          %d", failed)
    logging.info("Has features:    %d", has_features_count)
    logging.info("Exported SLDPRT: %d", exported)
    logging.info("")
    logging.info("Results: %s", results_excel_path)
    logging.info("Sidecar: %s", sidecar_path)

    # Verbose: dump all feature types encountered
    if args.verbose and global_all_types:
        logging.info("")
        logging.info("All feature types encountered:")
        for t in sorted(global_all_types):
            status = "allowed" if t in ALLOWED_FEATURE_TYPES else "FLAGGED"
            logging.info("  %-35s [%s]", t, status)

    # Save all-types to sidecar for auditing
    progress["__meta__"] = {
        "lastRun": datetime.now().isoformat(),
        "totalProcessed": processed,
        "succeeded": succeeded,
        "failed": failed,
        "hasFeatures": has_features_count,
        "exported": exported,
        "allFeatureTypesEncountered": sorted(global_all_types),
    }
    save_progress(sidecar_path, progress)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def setup_logging(output_dir, verbose=False):
    """Configure logging to both console and file."""
    os.makedirs(output_dir, exist_ok=True)
    log_file = os.path.join(
        output_dir,
        f"inspect_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log",
    )

    handlers = [
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(log_file, encoding="utf-8"),
    ]

    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)-5s %(message)s",
        datefmt="%H:%M:%S",
        handlers=handlers,
    )
    logging.info("Log file: %s", log_file)


def main():
    parser = argparse.ArgumentParser(
        description="Inspect SolidWorks assemblies for geometry features and export as SLDPRT",
    )
    parser.add_argument("-i", "--input", required=True,
                        help="Input Excel file (migration upload list)")
    parser.add_argument("-o", "--output", default="output/assembly_inspect",
                        help="Output directory (default: output/assembly_inspect)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Inspect only, no SolidWorks operations")
    parser.add_argument("--start-index", type=int, default=0,
                        help="Start at Nth assembly (0-based, for batching)")
    parser.add_argument("--count", type=int, default=-1,
                        help="Process at most N assemblies (-1 = all)")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="Log all feature types encountered (for allowlist auditing)")
    parser.add_argument("--force", action="store_true",
                        help="Re-process assemblies already in sidecar")
    parser.add_argument("--sidecar", default=None,
                        help="Path to progress sidecar JSON (default: <output>/inspect_progress.json)")

    args = parser.parse_args()

    # Validate input
    if not os.path.isfile(args.input):
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    setup_logging(args.output, args.verbose)
    process_assemblies(args)


if __name__ == "__main__":
    main()
